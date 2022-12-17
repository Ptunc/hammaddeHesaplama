from openpyxl import Workbook, load_workbook

def bilgiAlHesapla(cimento, sifirBes, besOniki, su, katki, hammaddeKodlari, hammaddeAdlari, baslangicSatiri):
    toplamHammadde = cimento + sifirBes + besOniki + su + katki
    kod = (input("Mamülün Stok Kodunu Girin: "))
    kilo = int(input("Mamülün Adet Başına Kilogramını Girin: "))
    if kilo >= toplamHammadde:
        cimento, sifirBes, besOniki, su, katki = cimento*2, sifirBes*2, besOniki*2, su*2, katki*2
        toplamHammadde = cimento + sifirBes + besOniki + su + katki
        oran = (toplamHammadde/kilo)*0.97
        oranliCimento, oranliSifirBes, oranliBesOniki, oranliSu, oranliKatki = cimento/oran, sifirBes/oran, besOniki/oran, su/oran, katki/oran
    else:
        oran = (toplamHammadde/kilo)*0.97
        oranliCimento, oranliSifirBes, oranliBesOniki, oranliSu, oranliKatki = float(cimento/oran), float(sifirBes/oran), float(besOniki/oran), float(su/oran), float(katki/oran)
        
    return kayitFonksiyonu(kod, oranliCimento, oranliSifirBes, oranliBesOniki, oranliSu, oranliKatki, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)

def kayitFonksiyonu(kod, oranliCimento, oranliSifirBes, oranliBesOniki, oranliSu, oranliKatki, hammaddeKodlari, hammaddeAdlari, baslangicSatiri):
    hammaddeler = [4]
    hammaddeler = [oranliCimento, oranliSifirBes, oranliBesOniki, oranliKatki]
    if(oranliKatki == 0): maddeAdeti = 3 
    else: maddeAdeti = 4
    for i in range (maddeAdeti):
        satir = i + baslangicSatiri
        ws.cell(satir,1).value = kod
        ws.cell(satir,6).value = i+1  # type: ignore 
        ws.cell(satir,7).value = hammaddeAdlari[i]
        ws.cell(satir,8).value = hammaddeKodlari[i]
        ws.cell(satir,9).value = "KG" # type: ignore 
        ws.cell(satir,10).value = hammaddeler[i]
    

    return baslangicSatiri + maddeAdeti

wb = load_workbook("deneme.xlsx")
ws = wb.active
kontrol = 1
hammaddeAdlari = [4]
hammaddeKodlari = [4]

hammaddeKodlari = ["C1", "İ05", "İ5-12", "KATKI0006"]
hammaddeAdlari = ["Çimento", "0-5 İscehisar", "5-12 İscehisar", "Nanoblock Katkı"]

print("------------------Mamül Reçete Otomasyonuna Hoşgeldiniz!----------------")
baslangicSatiri = int(input("Excel dosyasının kaçıncı satırından başlamak istiyorsunuz?: "))
while(kontrol):
    print("-----------------------------------------------------------------------")
    secim = int(input("Mamül Türü Seçin:\n 1. Boru\n 2. Baca Tabanı\n 3. Bilezik\n 4. Kapak\n 5. Parke\n 6. Bordür\n 7. Yağmur Oluğu\n 8. Parke İnce\n 9. Çıkış\n"))
    print("-----------------------------------------------------------------------")
    if secim == 1:
        subSecim = int(input(" 1. 150'lik Boru\n 2. 200'lük Boru\n 3. 300-400-500'lük Boru\n"))
        if subSecim == 1:
            baslangicSatiri = bilgiAlHesapla(275, 1522, 484, 139, 0, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)
        elif subSecim == 2:
            baslangicSatiri = bilgiAlHesapla(275, 1462, 545, 139, 0, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)
        elif subSecim == 3:
            baslangicSatiri = bilgiAlHesapla(275, 1542, 450, 139, 0, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)

    elif secim == 2:
        baslangicSatiri = bilgiAlHesapla(260, 1413, 610, 138, 0, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)

    elif secim == 3:
        baslangicSatiri = bilgiAlHesapla(260, 1352, 671, 138, 0, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)

    elif secim == 4:
        baslangicSatiri = bilgiAlHesapla(300, 840, 990, 190, 0, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)

    elif secim == 5:
        baslangicSatiri = bilgiAlHesapla(200, 962, 744, 65, 2.5, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)

    elif secim == 6:
        subSecim = int(input(" 1. 22'lik Bordür\n 2. 50'lik Bordür\n 3. 22'lik Bahçe Bordürü\n"))
        if subSecim == 1:
            baslangicSatiri = bilgiAlHesapla(184, 1200, 489, 65, 2.5, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)
        elif subSecim == 2:
            baslangicSatiri = bilgiAlHesapla(184, 966, 734, 65, 2.5, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)
        elif subSecim == 3:
            baslangicSatiri = bilgiAlHesapla(184, 1100, 594, 65, 2.5, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)

    elif secim == 7:
        baslangicSatiri = bilgiAlHesapla(184, 1100, 594, 65, 2.5, hammaddeKodlari, hammaddeAdlari, baslangicSatiri)


    elif secim == 9:
        print("Otomasyon Çalışmayı Durdurdu.")
        wb.save("denemenindenemesi.xlsx")
        kontrol = 0
    else: 
        print("Yanlış Giriş Yaptınız!")
        baslangicSatiri -=1
    baslangicSatiri +=1
